from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.drawing.image import Image
from openpyxl.utils import range_boundaries


# 创建一个空的工作簿
def create_excel(self):
    # 创建一个新的工作簿
    wb = Workbook()
    # 保存工作簿
    wb.save(self.excel_file)
    return wb


# 根据指定的单元格，单数据写入
def write_one(self, sheet_name, cells, value):
    # 打开Excel
    wb = self.open_excel
    # 指定Sheet
    ws = self.sheet(wb, sheet_name)
    # 写入
    ws[cells] = value
    # 保存
    self.save_excel(wb)
    return f"成功在[{sheet_name}]工作表，{cells}单元格插入1条数据"


# 根据指定的单元格，多数据写入
def write_many(self, sheet_name, data: dict):
    wb = self.open_excel
    ws = self.sheet(wb, sheet_name)
    for k, v in data.items():
        ws[k] = v
    self.save_excel(wb)
    count = data.__len__()
    return f"成功在[{sheet_name}]工作表插入{count}条数据"


# 多数据批量写入工作表中，此方式不需要指定单元格，默认从A1开始写入
def write_batch(self, sheet_name, data: list):
    wb = self.open_excel
    ws = self.sheet(wb, sheet_name)
    for row in data:
        for i in range(len(row)):
            ws.cell(row=data.index(row) + 1, column=i + 1, value=row[i])
    self.save_excel(wb)
    count = len(data)
    return f"成功在[{sheet_name}]工作表插入{count}条数据"


# 删除指定单元格的数据，单数据删除
def delete_one(self, sheet_name, cells):
    wb = self.open_excel
    ws = wb[sheet_name]
    del ws[cells]
    self.save_excel(wb)
    return f"成功在[{sheet_name}]工作表，删除{cells}单元格的数据"


# 删除指定单元格的数据，多数据删除
def delete_many(self, sheet_name, cells: list):
    wb = self.open_excel
    ws = wb[sheet_name]
    for cell in cells:
        del ws[cell]
    count = len(cells)
    self.save_excel(wb)
    return f"成功在[{sheet_name}]工作表删除{count}条数据"


# 删除指定工作表
def delete_sheet(self, sheet_name):
    wb = self.open_excel
    ws = wb[sheet_name]
    wb.remove(ws)
    self.save_excel(wb)
    return f"成功删除工作表[{sheet_name}]"


# 读取工作表中，指定单元格的内容
def read_one(self, sheet_name, cells):
    wb = self.open_excel
    ws = wb[sheet_name]
    value = ws[cells].value
    return value


# 读取工作表中，指定多个单元格的内容
def read_many(self, sheet_name, cells: list):
    wb = self.open_excel
    ws = wb[sheet_name]
    values = []
    for cell in cells:
        value = ws[cell].value
        values.append(value)
    return values


# 读取指定工作表的所有内容
def read_all(self, sheet_name):
    wb = self.open_excel
    ws = wb[sheet_name]
    all_values = []
    for row in ws.iter_rows():
        row_values = []
        for cell in row:
            row_values.append(cell.value)
        all_values.append(row_values)
    return all_values


# 读取Excel文档中所有工作表的名称
def read_sheets(self):
    wb = self.open_excel
    sheet_names = wb.sheetnames
    return sheet_names


# 根据工作表中的数据，生成折线图
def line_chart(self, sheet_name, position, title):
    wb = self.open_excel
    ws = self.sheet(wb, sheet_name)
    # 创建一个折线图
    chart = LineChart()
    # 获取数据源范围
    calculate_dimension_source = ws.calculate_dimension()
    min_row, min_col, max_row, max_col = range_boundaries(calculate_dimension_source)
    min_col += 1
    range_data_ = "{}{}:{}{}".format(ws.cell(min_row, min_col).column_letter, min_row,
                                     ws.cell(max_col, max_row).column_letter, max_col)
    range_data = "{}!{}".format(ws.title, range_data_)
    data_ = Reference(ws, range_string=range_data)
    # 将数据源添加到折线图上
    chart.add_data(data_, titles_from_data=True)
    chart.title = title
    # 设置 x 轴的数据源
    min_row += 1
    min_col -= 1
    range_string1_ = "{}{}:{}{}".format(ws.cell(min_row, min_col).column_letter, min_row,
                                        ws.cell(min_row, min_col).column_letter, max_col)
    range_string1 = "{}!{}".format(ws.title, range_string1_)
    x_axis_data = Reference(ws, range_string=range_string1)
    chart.set_categories(x_axis_data)
    # 将折线图插入到工作表中
    ws.add_chart(chart, position)
    self.save_excel(wb)
    return f"成功在工作表[{sheet_name}]中生成折线图！"


# 在工作表中指定的单元格位置插入宽为400，高为200的图片
def insert_image(self, sheet_name, position, image, width=400, height=200):
    wb = self.open_excel
    ws = self.sheet(wb, sheet_name)
    # 创建Image对象
    img = Image(image)
    # 设置图片大小
    img.width = width
    img.height = height
    # 将Image对象插入到工作表中指定位置的单元格中
    ws.add_image(img, position)
    self.save_excel(wb)
    return f"成功在[{sheet_name}]工作表插入图片{image}"


# 提取工作表中的图片，并保存到get_images文件夹中
def get_images(self, sheet_name):
    wb = self.open_excel
    ws = wb[sheet_name]
    images = ws._images
    for index, img in enumerate(images):
        data = img._data()
        with open(f"get_images/image{str(index)}.png", "wb") as img:
            img.write(data)
    return f"成功提取出工作表[{sheet_name}]中的图片！"


'''
import xlwt
def crate_excel():
    book_name = input('输入工作簿名字:')
    sheet_name = input('输入工作表名字:')
    workbook = xlwt.Workbook(encoding="utf-8")
    worksheet = workbook.add_sheet(sheet_name)
    workbook.save(book_name + '.xls')
if __name__ == '__main__':
    crate_excel()
'''


# create_excel(1)


# class test1:
#     def __init__(self):
#


x = lambda a: a + 10

print(x(2))